import os

import openpyxl

from tests.conftest import PROJECT_DIR


class HomePageData:
    test_HomePage_data = [{"firstname": "Kiran", "lastname": "Bellad", "gender": "Male"},
                          {"firstname": "Raghavendra", "lastname": "Bellad", "gender": "Male"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook(PROJECT_DIR + "\\TestData\\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    # Dict["lastname"]="shetty
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]
